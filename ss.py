
# --- Optional runtime bootstrap (fallback only) ---
# Tries to install openpyxl at runtime if it's missing.
# Not recommended for Streamlit Cloud, but handy for local runs without requirements setup.

def _ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
        return
    except Exception:
        pass

    import sys, subprocess
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "--no-cache-dir", "openpyxl>=3.1"])
        import openpyxl  # noqa: F401
    except Exception as e:
        import streamlit as st
        st.error(
            "Failed to install **openpyxl** at runtime. "
            "Please ensure `requirements.txt` includes `openpyxl>=3.1` and redeploy."
        )
        st.stop()

_ensure_openpyxl()
# --- End fallback ---




# app.py
# -*- coding: utf-8 -*-
"""
Packing List Recalculator (Gas Cookers) – Order-driven ratios

Workflow (tabs):
1) Order (master): paste a 3-column table -> Material code, Name description, Full quantity of the order
2) Packing Lists: upload the workbook (.xlsx) with any number of sheets/tables
3) Options & Run: enter Order cooker quantity, Target cooker quantity (for these packing lists), run
4) Preview & Download: inspect results and export the recalculated workbook

Key design:
- GROUP BY CODE ONLY (SAP material code). Names never used for grouping.
- Safe READ/WRITE for merged cells (always use master top-left cell).
- Only update numeric Qu. cells; text/unit rows remain intact.
- Keep structure & styles: same sheets, same boxes/rows; only Qu. values change.
"""

import io
import re
from typing import Dict, List, Tuple, Any, Optional

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# -------------------------
# Page setup
# -------------------------
st.set_page_config(page_title="Packing List Recalculator (Order-driven)", layout="wide")
st.title("Packing List Recalculator (Order-driven ratios)")

# -------------------------
# Helpers & constants
# -------------------------
EXPECTED_HEADERS = {
    "S.N", "Box code", "component in arabic", "component in E", "Codes", "Qu.", "Box"
}

def norm(x: Any) -> str:
    if x is None:
        return ""
    return str(x).strip()

def parse_quantity(q_raw: Any) -> Tuple[bool, int]:
    """
    Extract a numeric quantity from a cell value.
    Returns (is_numeric, int_value).
    - Accepts numbers stored as text/float ('120', '120.0').
    - If text contains a number with words ('4 بالته'), extracts the first number.
    - If nothing numeric is found, returns (False, 0).
    """
    s = norm(q_raw)
    if s == "":
        return (False, 0)
    try:
        val = float(s)
        return (True, int(round(val)))
    except Exception:
        pass
    m = re.search(r"[-+]?\d+(?:\.\d+)?", s)
    if m:
        try:
            val = float(m.group(0))
            return (True, int(round(val)))
        except Exception:
            return (False, 0)
    return (False, 0)

def is_header_row(values: List[Any]) -> bool:
    vals = {norm(v) for v in values if norm(v)}
    essential = {"S.N", "Codes", "Qu."}
    return essential.issubset(vals)

# -------------------------
# merged cell safe read & write
# -------------------------
def find_merged_master(ws: Worksheet, row: int, col: int) -> Optional[Tuple[int, int]]:
    coord = ws.cell(row=row, column=col).coordinate
    for rng in ws.merged_cells.ranges:
        if coord in rng:
            return (rng.min_row, rng.min_col)
    return None

def safe_read(ws: Worksheet, row: int, col: int) -> Any:
    master = find_merged_master(ws, row, col)
    if master is not None:
        return ws.cell(row=master[0], column=master[1]).value
    return ws.cell(row=row, column=col).value

def safe_write(ws: Worksheet, row: int, col: int, value: int) -> None:
    master = find_merged_master(ws, row, col)
    target_row, target_col = (master if master is not None else (row, col))
    ws.cell(row=target_row, column=target_col).value = int(value)

# -------------------------
# table detection
# -------------------------
def find_tables(ws: Worksheet) -> List[Dict[str, Any]]:
    """
    Find tables by header rows. Return: [{header_row, col_map, data_rows}]
    """
    tables = []
    max_row = ws.max_row
    max_col = ws.max_column
    row_idx = 1
    while row_idx <= max_row:
        row_vals = [safe_read(ws, row_idx, c) for c in range(1, max_col + 1)]
        if is_header_row(row_vals):
            col_map = {}
            for c in range(1, max_col + 1):
                h = norm(safe_read(ws, row_idx, c))
                if h in EXPECTED_HEADERS:
                    col_map[h] = c
            data_rows = []
            r = row_idx + 1
            while r <= max_row:
                candidate = [safe_read(ws, r, c) for c in range(1, max_col + 1)]
                if is_header_row(candidate):
                    break
                data_rows.append(r)
                r += 1
            tables.append({"header_row": row_idx, "col_map": col_map, "data_rows": data_rows})
            row_idx = r
        else:
            row_idx += 1
    return tables

# -------------------------
# occurrences (group by CODE ONLY)
# -------------------------
# Occurrence tuple:
# (ws_name, row_idx, q_col, numeric_q, is_numeric, arabic_name, code)
Occurrence = Tuple[str, int, int, int, bool, str, str]

def component_id_from_row(ws: Worksheet, row: int, col_map: Dict[str, int]) -> str:
    code = norm(safe_read(ws, row, col_map.get("Codes", 0)))
    if code:
        return code
    # No code => keep row isolated; don't merge with other no‑code rows
    return f"__NO_CODE__@{ws.title}@{row}"

def extract_component_occurrences(wb) -> Dict[str, List[Occurrence]]:
    comp_occ: Dict[str, List[Occurrence]] = {}
    for ws in wb.worksheets:
        tables = find_tables(ws)
        for tbl in tables:
            cmap = tbl["col_map"]
            q_col = cmap.get("Qu.")
            if not q_col:
                continue
            for r in tbl["data_rows"]:
                all_vals = [safe_read(ws, r, c) for c in range(1, ws.max_column + 1)]
                if not any(norm(v) for v in all_vals):
                    continue
                q_raw = safe_read(ws, r, q_col)
                is_num, q_val = parse_quantity(q_raw)
                arabic_name = norm(safe_read(ws, r, cmap.get("component in arabic", 0)))
                cid = component_id_from_row(ws, r, cmap)
                code_display = cid if not cid.startswith("__NO_CODE__@") else ""
                comp_occ.setdefault(cid, []).append((ws.title, r, q_col, q_val, is_num, arabic_name, code_display))
    return comp_occ

# -------------------------
# allocation
# -------------------------
def largest_remainder_allocate(originals: List[int], target_total: int) -> List[int]:
    if target_total < 0:
        target_total = 0
    n = len(originals)
    if n == 0:
        return []
    s = sum(originals)
    if s == 0:
        base = target_total // n
        rem = target_total - base * n
        alloc = [base] * n
        for i in range(rem):
            alloc[i] += 1
        return alloc
    weights = np.array(originals, dtype=float) / float(s)
    ideal = weights * float(target_total)
    floors = np.floor(ideal).astype(int)
    residual = int(target_total - floors.sum())
    fracs = ideal - floors
    order = np.argsort(-fracs)
    alloc = floors.copy()
    for i in range(residual):
        alloc[order[i]] += 1
    return alloc.tolist()

def apply_allocations(wb, comp_occ: Dict[str, List[Occurrence]], comp_targets: Dict[str, int]) -> None:
    ws_index = {ws.title: ws for ws in wb.worksheets}
    for cid, occs in comp_occ.items():
        target_total = comp_targets.get(cid)
        if target_total is None:
            continue
        numeric_occs = [(ws, r, c, q, is_num, ar, code) for (ws, r, c, q, is_num, ar, code) in occs if is_num]
        if not numeric_occs:
            continue
        originals = [q for (_, _, _, q, _, _, _) in numeric_occs]
        new_vals = largest_remainder_allocate(originals, target_total)
        for new_q, (ws_name, row_idx, q_col, _, _, _, _) in zip(new_vals, numeric_occs):
            ws = ws_index[ws_name]
            safe_write(ws, row_idx, q_col, int(new_q))

# -------------------------
# ratios & targets (ORDER-driven)
# -------------------------
def compute_order_ratios(order_df: pd.DataFrame, order_cookers: int) -> Dict[str, float]:
    """
    order_df columns (required):
      - Material code (string)
      - Name description (string) [not used for grouping]
      - Full quantity of the order (numeric)
    Returns: dict[code] -> per-cooker ratio
    """
    ratios: Dict[str, float] = {}
    if order_cookers <= 0 or order_df.empty:
        return ratios
    # Clean columns
    df = order_df.copy()
    df.columns = [c.strip().lower() for c in df.columns]
    # Expected normalized names
    # 'material code', 'name description', 'full quantity of the order'
    # Be robust to slight naming differences:
    col_code = next((c for c in df.columns if "material" in c and "code" in c), "material code")
    col_qty = next((c for c in df.columns if "quantity" in c and "order" in c), "full quantity of the order")
    if col_code not in df.columns or col_qty not in df.columns:
        return ratios
    # Drop blanks and non-numeric quantities
    df = df[[col_code, col_qty]].copy()
    df[col_code] = df[col_code].astype(str).str.strip()
    def _to_int(x):
        try:
            return int(round(float(str(x).strip())))
        except Exception:
            return 0
    df[col_qty] = df[col_qty].map(_to_int)
    df = df[df[col_code] != ""]
    df_grp = df.groupby(col_code, as_index=True)[col_qty].sum()
    for code, total in df_grp.items():
        ratios[code] = total / float(order_cookers)
    return ratios

def build_targets_from_order_ratios(order_ratios: Dict[str, float], target_cookers: int,
                                    workbook_codes: List[str]) -> Dict[str, int]:
    """
    For each code that appears in the workbook, if it exists in the order ratios,
    build target_total = round(ratio * target_cookers). If missing, leave None (unchanged).
    """
    targets: Dict[str, int] = {}
    for cid in workbook_codes:
        if cid.startswith("__NO_CODE__@"):
            # No-code rows left unchanged
            continue
        r = order_ratios.get(cid)
        if r is None:
            # Missing code in order => keep as-is (target not set)
            continue
        targets[cid] = max(int(round(r * float(target_cookers))), 0)
    return targets

# -------------------------
# UI tabs
# -------------------------
tab_order, tab_packing, tab_run, tab_preview = st.tabs(
    ["1) Order (master)", "2) Packing Lists", "3) Options & Run", "4) Preview & Download"]
)

with tab_order:
    st.subheader("Paste the full order (master shipment)")
    st.write("Required columns (in this order): **Material code**, **Name description**, **Full quantity of the order**")
    example = pd.DataFrame({
        "Material code": ["100001234", "400045501"],
        "Name description": ["Fan motor", "Carton Plaza 60"],
        "Full quantity of the order": [2500, 600]
    })
    st.caption("Tip: Click into the table below and paste your rows (Ctrl+V). You can also load from CSV/Excel and paste here.")
    order_df = st.data_editor(
        example,
        use_container_width=True,
        num_rows="dynamic",
        key="order_data_editor",
    )

with tab_packing:
    st.subheader("Upload the packing lists workbook (.xlsx)")
    st.write("The app will handle multiple sheets and multiple table sections per sheet.")
    uploaded = st.file_uploader("Upload Excel file", type=["xlsx"], key="packing_uploader")
    if uploaded is not None:
        data = uploaded.read()
        bio_in_tmp = io.BytesIO(data)
        wb_tmp = load_workbook(filename=bio_in_tmp, data_only=True)
        st.write(f"Detected {len(wb_tmp.worksheets)} worksheet(s).")
        for ws in wb_tmp.worksheets:
            st.write(f"• **{ws.title}**")
            # lightweight table count preview
            tables = find_tables(ws)
            st.write(f"   Tables detected: {len(tables)}")

with tab_run:
    st.subheader("Options & Run")
    colA, colB = st.columns(2)
    with colA:
        order_cookers = st.number_input("Order cooker quantity (for ratio calculation)", min_value=1, value=250, step=1)
    with colB:
        target_cookers = st.number_input("Target cooker quantity (for these packing lists)", min_value=1, value=250, step=1)

    st.caption("Ratios are computed from the order table, not from packing lists. We'll scale packing list quantities to match target cookers.")
    exclude_packaging = st.checkbox("Keep rows without code (and any text/unit quantities) unchanged", value=True,
                                    help="Rows missing a material code or with non-numeric Qu. will not be modified.")

    run_btn = st.button("Compute & Prepare Output", type="primary")

    if run_btn:
        # Validate order df
        df_order = order_df.copy()
        df_order.columns = [c.strip() for c in df_order.columns]
        needed_cols = {"Material code", "Name description", "Full quantity of the order"}
        if not needed_cols.issubset(set(df_order.columns)):
            st.error("Order table must have columns: Material code, Name description, Full quantity of the order")
        elif uploaded is None:
            st.error("Please upload the packing lists workbook in tab 2.")
        else:
            # Compute order ratios
            order_ratios = compute_order_ratios(df_order, order_cookers)
            if not order_ratios:
                st.error("No valid order ratios could be computed. Check 'Material code' and 'Full quantity of the order' columns and 'Order cooker quantity'.")
            else:
                # Load workbook (twice: once for analysis, once for writing with styles)
                raw = uploaded.getvalue()
                bio_in = io.BytesIO(raw)
                wb_in = load_workbook(filename=bio_in, data_only=True)
                comp_occ = extract_component_occurrences(wb_in)

                # Gather workbook codes
                workbook_codes = list(comp_occ.keys())

                # Build targets for workbook codes using order ratios
                targets = build_targets_from_order_ratios(order_ratios, target_cookers, workbook_codes)

                # Prepare a session buffer for output
                bio_in.seek(0)
                wb_out = load_workbook(filename=io.BytesIO(bio_in.read()), data_only=False)

                # Apply allocations (numeric-only; no-code rows left unchanged)
                apply_allocations(wb_out, comp_occ, targets)

                # Save in session state for preview/download
                out_buf = io.BytesIO()
                wb_out.save(out_buf)
                out_buf.seek(0)
                st.session_state["out_buf"] = out_buf
                st.session_state["comp_occ"] = comp_occ
                st.session_state["targets"] = targets
                st.session_state["order_ratios"] = order_ratios
                st.success("Output prepared. Go to '4) Preview & Download' to inspect and export.")

with tab_preview:
    st.subheader("Preview & Download")
    if "comp_occ" in st.session_state and "targets" in st.session_state:
        comp_occ = st.session_state["comp_occ"]
        targets = st.session_state["targets"]
        order_ratios = st.session_state.get("order_ratios", {})

        # Build preview DF (Arabic names + code)
        rows = []
        for cid, occs in comp_occ.items():
            code_display = cid if not cid.startswith("__NO_CODE__@") else ""
            arabic_name = occs[0][5] if occs else ""
            original_total_numeric = sum(q for (_, _, _, q, is_num, _, _) in occs if is_num)
            ratio = order_ratios.get(cid, None)
            tgt = targets.get(cid, None)
            rows.append({
                "Code": code_display,
                "Component (Arabic)": arabic_name,
                "Original total (numeric Qu.)": original_total_numeric,
                "Per-cooker ratio (from order)": (ratio if ratio is not None else ""),
                "Target total (Qu.)": (tgt if tgt is not None else ""),
                "Occurrences (numeric/all)": f"{sum(1 for o in occs if o[4])}/{len(occs)}"
            })
        df_prev = pd.DataFrame(rows).sort_values(["Code", "Component (Arabic)"]).reset_index(drop=True)
        st.dataframe(df_prev, use_container_width=True)

        if "out_buf" in st.session_state:
            st.download_button(
                label="Download recalculated packing list (.xlsx)",
                data=st.session_state["out_buf"],
                file_name="packing_list_recalculated.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                       )



# ==== Centered footer ====
footer_css = """
<style>
.app-footer {
  position: fixed;
  left: 50%;
  bottom: 12px;
  transform: translateX(-50%);
  z-index: 9999;
  background: rgba(255,255,255,0.85);
  border: 1px solid #e6e6e6;
  border-radius: 14px;
  padding: 8px 14px;
  font-weight: 600;
  font-size: 14px;
}
</style>
"""
footer_html = """
<div class="app-footer">✨ تم التنفيذ بواسطة م / بيتر عمانوئيل – جميع الحقوق محفوظة © 2025 ✨</div>
"""
st.markdown(footer_css, unsafe_allow_html=True)
st.markdown(footer_html, unsafe_allow_html=True)



